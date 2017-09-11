#!python3
# -*- coding: utf-8 -*-
# requed:
# https://openpyxl.readthedocs.io/en/default/
# 
import argparse
import os, os.path
from openpyxl import Workbook, load_workbook
import sys


# 格式化性别 男性 > 男 , 女性 > 女
def formatSex(s):
    if s != r'男性' and s != r'女性':
        raise NameError('性别错误')
    return '男' if s == r'男性' else '女'

# 格式化日期  20110817 > 2011-08-17
def formatDate(s):
    s = str(s)
    return '%s-%s-%s' % (s[0:4], s[4:6], s[6:8])

def formatNo(s):
    return s

# 复制的字段信息
# 复制 {fromName} 列数据到 {toName} 列
# 使用 {formatter} 函数格式化数据
class CP:
    def __init__(self, fromName, toName, formatter = None, isPutBySelf = False):
        self.fromName = fromName
        self.toName = toName
        self.formatter = formatter if formatter is not None else formatNo
        self.isPutBySelf = isPutBySelf

# 复制的列信息
cps = [
    CP(r'学生姓名', r'姓名', None, True),
    CP(r'学生身份证件号', r'身份证编号'),
    CP(r'学生性别', r'性别', formatSex),
    CP(r'学生出生日期', r'出生日期', formatDate),
    CP(r'学生户口详细地址', r'户籍地址'),
    CP(r'家庭成员1家庭成员1证件号码', r'居住证编号'),
    CP(r'家庭成员1家庭成员1居住证有效期', r'居住证有效期'),
    CP(r'学生现住址详细地址', r'联系地址'),
    CP(r'家庭成员1家庭成员1联系电话', r'联系电话'),
]


# 过滤from表中的数据 (colName, value) -> bool
# 只有通过的数据需要复制
# dataFilter = lambda colName, value: True
dataFilter = lambda colName, value: colName != r'学生班级名称' or value == r'一年级2班'


putBySelfIdx = -1
isPutBySelfs = []
for i, cp in enumerate(cps):
    if cp.isPutBySelf:
        isPutBySelfs.append(cp)
        putBySelfIdx = i
if len(isPutBySelfs) > 1:
    raise NameError('isPutBySelf must be less 2')

# print(putBySelfIdx, isPutBySelfs)

def readTitleMap(ws):
    col = ws.max_column
    row = ws.max_row
    titleMap = {}
    for i in range(1, ws.max_column + 1):
        titleMap[ws.cell(row = 1, column = i).value] = i
    return titleMap

def writeexcel(excelpath, datas):
    wb = load_workbook(os.path.abspath(excelpath))
    ws = wb.active
    col = ws.max_column
    row = ws.max_row

    titleMap = readTitleMap(ws)
    colIdxs = []
    for cp in cps:
        colIdxs.append(titleMap[cp.toName])

    nameMap = None

    if putBySelfIdx >= 0:
        nameColIdx = colIdxs[putBySelfIdx]
        nameMap = {}
        for i in range(2, row + 1):
            name = ws.cell(row = i, column = nameColIdx).value
            nameMap[name] = i


    for i, data in enumerate(datas):
        # print(dname)
        row = i + 2
        if putBySelfIdx >= 0:
            row = nameMap[data[putBySelfIdx]]
        for ii, cp in enumerate(cps):
            v = cp.formatter(data[ii])
            ws.cell(row = row, column = colIdxs[ii], value = v)

    wb.save(os.path.abspath(excelpath))

def vieryDataFilter(ws, row, titleMap):
    for colName, col in titleMap.items():
        v = ws.cell(row = row, column = col).value
        if not dataFilter(colName, v):
            return False
    return True

def readexcel(excelpath):
    print(excelpath)
    wb = load_workbook(os.path.abspath(excelpath))
    ws = wb.active
    col = ws.max_column
    row = ws.max_row
    if row <= 1:
        print('!!!!!empty excel!!!!!')
        return None
    # print(col, row)
    # ws.Columns.AutoFit()
    titleMap = readTitleMap(ws)

    colIdxs = []
    for cp in cps:
        colIdxs.append(titleMap[cp.fromName])

    datas = []
    for i in range(2, row + 1):
        if vieryDataFilter(ws, i, titleMap):
            one = []
            for j in colIdxs:
                v = ws.cell(row = i, column = j).value
                one.append(v)
            datas.append(one)

    return datas

# -------------- main ----------------
if __name__ == '__main__':
    parser = argparse.ArgumentParser(usage='%(prog)s <inputExcel> <outputExcel>',
        description='copy $inputExcel data to $outExcel',
        formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('inputExcel',
        help='input excel filename')
    parser.add_argument('outputExcel',
        help='output excel filename')

    args = parser.parse_args()

    datas = readexcel(args.inputExcel)
    writeexcel(args.outputExcel, datas)

    # datas = readexcel('data.xlsx')
    # writeexcel('out.xlsx', datas)
    
